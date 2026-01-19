from typing import Any, Optional
from openpyxl import load_workbook
import xlwings as xw
from ..setup import get_current_instance


def get_excel_table(
    excel_path: str,
    range: str,
    values: dict[str, Any] | None = None,
    cache: bool = True,
) -> str:
    """
    截取 Excel 表格的指定范围并转换为 HTML

    Args:
        excel_path: Excel 文件路径
        values: 要更新的值，字典格式 {'A1': value1, 'Sheet1!B2': value2}
        screenshot_range: 要截取的范围，支持 'A1:E10' 或 'Sheet1!A1:E10' 格式

    Returns:
        HTML table 字符串

    Example:
        >>> html = screenshot_sheet(
        ...     'data.xlsx',
        ...     {'Sheet1!A1': 100, 'Sheet2!A2': 200},
        ...     'Sheet2!A1:C10'
        ... )
    """

    ctx = get_current_instance()
    json_db = ctx.get_json_db()

    key_obj = (excel_path, range, values)
    if cache:
        # 判断是否有缓存
        cached_html = json_db.get(key_obj, None)
        if cached_html is not None:
            return cached_html

    # 使用 xlwings 更新值并计算公式
    if values:
        app = xw.App(visible=False)
        try:
            wb = app.books.open(excel_path)
            # 更新值
            for cell_address, value in values.items():
                # 解析单元格地址，支持 'Sheet1!A1' 格式
                if "!" in cell_address:
                    sheet_name_val, cell_ref = cell_address.split("!", 1)
                    wb.sheets[sheet_name_val].range(cell_ref).value = value
                else:
                    wb.sheets[0].range(cell_address).value = value
            # 保存并关闭（会自动计算公式）
            wb.save()
            wb.close()
        finally:
            app.quit()

    # 使用 openpyxl 读取计算后的值并转换为 HTML
    processor = ExcelProcessor(excel_path)
    try:
        # 解析导出范围，提取工作表名称
        sheet_name = None
        range_address = range
        if "!" in range:
            sheet_name, range_address = range.split("!", 1)

        processor.load(sheet_name=sheet_name, data_only=True)
        html = processor.range_to_html(range_address, include_styles=True)

        # 将结果保存到缓存
        if cache:
            json_db.set(key_obj, html)

        return html
    finally:
        processor.close()


class ExcelProcessor:
    """
    Excel 处理器，用于读取 Excel 并转换为 HTML table，支持合并单元格和样式
    """

    def __init__(self, excel_path: str):
        """
        初始化 Excel 处理器

        Args:
            excel_path: Excel 文件路径
        """
        self.excel_path = excel_path
        self.workbook = None
        self.worksheet = None

    def load(self, sheet_name: Optional[str] = None, data_only: bool = True):
        """
        加载 Excel 文件

        Args:
            sheet_name: 工作表名称，如果为 None 则使用第一个工作表
            data_only: 如果为 True，则读取公式的计算值；如果为 False，则读取公式本身
        """
        self.workbook = load_workbook(self.excel_path, data_only=data_only)
        if sheet_name:
            self.worksheet = self.workbook[sheet_name]
        else:
            self.worksheet = self.workbook.active
        return self

    def get_merged_cells_map(self):
        """
        获取合并单元格的映射

        Returns:
            字典，键为合并单元格中的每个单元格坐标，值为合并区域的主单元格坐标
        """
        if not self.worksheet:
            raise ValueError("请先调用 load() 方法加载工作表")

        merged_map = {}
        for merged_range in self.worksheet.merged_cells.ranges:
            # 获取合并区域的主单元格（左上角单元格）
            min_row, min_col = merged_range.min_row, merged_range.min_col
            main_cell = (min_row, min_col)

            # 将合并区域内的所有单元格都映射到主单元格
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_map[(row, col)] = {
                        "main_cell": main_cell,
                        "rowspan": merged_range.max_row - merged_range.min_row + 1,
                        "colspan": merged_range.max_col - merged_range.min_col + 1,
                        "is_main": (row == min_row and col == min_col),
                    }

        return merged_map

    def range_to_html(self, range_address: str, include_styles: bool = True) -> str:
        """
        将指定范围转换为 HTML table

        Args:
            range_address: 范围地址，如 'A1:E10'
            include_styles: 是否包含样式（边框、对齐等）

        Returns:
            HTML table 字符串
        """
        if not self.worksheet:
            raise ValueError("请先调用 load() 方法加载工作表")

        # 解析范围
        cell_range = self.worksheet[range_address]

        # 获取合并单元格映射
        merged_map = self.get_merged_cells_map()

        # 开始构建 HTML
        html_parts = ['<table border="1" style="border-collapse: collapse;">']

        # 遍历范围内的行
        for row_idx, row in enumerate(cell_range):
            html_parts.append("  <tr>")

            # 遍历行中的单元格
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue

                cell_coord = (cell.row, cell.column)

                # 检查是否是合并单元格
                if cell_coord in merged_map:
                    merge_info = merged_map[cell_coord]

                    # 如果不是主单元格，跳过（已由主单元格处理）
                    if not merge_info["is_main"]:
                        continue

                    # 如果是主单元格，添加 rowspan 和 colspan
                    rowspan = merge_info["rowspan"]
                    colspan = merge_info["colspan"]

                    style_parts = []
                    if include_styles:
                        # 添加对齐方式
                        if cell.alignment:
                            if cell.alignment.horizontal:
                                style_parts.append(
                                    f"text-align: {cell.alignment.horizontal}"
                                )
                            if cell.alignment.vertical:
                                style_parts.append(
                                    f"vertical-align: {cell.alignment.vertical}"
                                )

                        # 添加字体样式
                        if cell.font:
                            if cell.font.bold:
                                style_parts.append("font-weight: bold")
                            if cell.font.italic:
                                style_parts.append("font-style: italic")
                            if cell.font.color and cell.font.color.rgb:
                                rgb = str(cell.font.color.rgb)
                                style_parts.append(f"color: #{rgb[2:]}")

                        # 添加背景色
                        if (
                            cell.fill
                            and cell.fill.start_color
                            and cell.fill.start_color.rgb
                        ):
                            rgb = str(cell.fill.start_color.rgb)
                            if rgb != "00000000":  # 不是默认颜色
                                style_parts.append(f"background-color: #{rgb[2:]}")

                    style_attr = (
                        f' style="{"; ".join(style_parts)}"' if style_parts else ""
                    )
                    rowspan_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ""
                    colspan_attr = f' colspan="{colspan}"' if colspan > 1 else ""

                    value = cell.value if cell.value is not None else ""
                    html_parts.append(
                        f"    <td{rowspan_attr}{colspan_attr}{style_attr}>{value}</td>"
                    )
                else:
                    # 普通单元格
                    style_parts = []
                    if include_styles:
                        # 添加对齐方式
                        if cell.alignment:
                            if cell.alignment.horizontal:
                                style_parts.append(
                                    f"text-align: {cell.alignment.horizontal}"
                                )
                            if cell.alignment.vertical:
                                style_parts.append(
                                    f"vertical-align: {cell.alignment.vertical}"
                                )

                        # 添加字体样式
                        if cell.font:
                            if cell.font.bold:
                                style_parts.append("font-weight: bold")
                            if cell.font.italic:
                                style_parts.append("font-style: italic")
                            if cell.font.color and cell.font.color.rgb:
                                rgb = str(cell.font.color.rgb)
                                style_parts.append(f"color: #{rgb[2:]}")

                        # 添加背景色
                        if (
                            cell.fill
                            and cell.fill.start_color
                            and cell.fill.start_color.rgb
                        ):
                            rgb = str(cell.fill.start_color.rgb)
                            if rgb != "00000000":  # 不是默认颜色
                                style_parts.append(f"background-color: #{rgb[2:]}")

                    style_attr = (
                        f' style="{"; ".join(style_parts)}"' if style_parts else ""
                    )
                    value = cell.value if cell.value is not None else ""
                    html_parts.append(f"    <td{style_attr}>{value}</td>")

            html_parts.append("  </tr>")

        html_parts.append("</table>")
        return "\n".join(html_parts)

    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None
